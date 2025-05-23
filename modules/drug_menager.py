import customtkinter as ctk
from openpyxl import load_workbook
import os

class drugs(ctk.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.title("Przeglądaj leki")
        self.geometry("700x600")
        self.configure(fg_color="#1f2937")
        self.focus()
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="Lista leków",
            font=("Arial", 34, "bold"),
            text_color="#329e76"
        ).pack(pady=(20, 10))

        self.scroll_frame = ctk.CTkScrollableFrame(self, fg_color="#1f2937", width=650)
        self.scroll_frame.pack(fill="both", expand=True, padx=20, pady=10)

        self.load_drugs_from_excel()

    def load_drugs_from_excel(self):
        filepath = os.path.join("database", "drugs.xlsx")
        workbook = load_workbook(filepath)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            id, nazwa, cena, stan_magazynowy = row

            card = ctk.CTkFrame(self.scroll_frame, fg_color="#2c3e50")
            card.pack(pady=10, fill="x", padx=10)

            ctk.CTkLabel(card, text=nazwa, font=("Arial", 20, "bold"), text_color="#329e76").pack(anchor="w", padx=10, pady=(5, 0))
            ctk.CTkLabel(card, text=f"ID: {id}", font=("Arial", 16), text_color="white").pack(anchor="w", padx=10)
            ctk.CTkLabel(card, text=f"Cena: {cena} zł", font=("Arial", 16), text_color="white").pack(anchor="w", padx=10)
            ctk.CTkLabel(card, text=f"Stan magazynowy: {int(stan_magazynowy)}", font=("Arial", 16), text_color="white").pack(anchor="w", padx=10, pady=(0, 5))