from openpyxl import load_workbook, Workbook
import os

def add_drug(name, price, stock=0, prescription_number="0"):
    filepath = "database/drugs.xlsx"

    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.append(["ID", "nazwa", "cena", "stan_magazynowy", "numer_recepty"])
    else:
        wb = load_workbook(filepath)
        ws = wb.active

        headers = [cell.value for cell in ws[1]] # pobieranie nagłówków z pierwszego wiersza
        if "numer_recepty" not in headers:
            ws.insert_cols(len(headers) + 1) # dodaje nowa kolumne na koncu arkusza
            ws.cell(row=1, column=len(headers) + 1, value="numer_recepty") #wypisuje nagłówek

    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True): #przechodzi odrazu do 2 wiersza
        if row[0] and isinstance(row[0], int): #sprawdza id w pierwszej kolumnie
            max_id = max(max_id, row[0])

    new_id = max_id + 1

    ws.append([new_id, name, price, stock, prescription_number])
    wb.save(filepath)


def remove_drug(identifier):
    filepath = "database/drugs.xlsx"

    if not os.path.exists(filepath):
        return False

    wb = load_workbook(filepath)
    ws = wb.active
    removed = False

    rows = list(ws.iter_rows(min_row=2))

    for row in rows:
        id_cell = row[0].value
        name_cell = row[1].value

        if str(identifier) == str(id_cell) or str(identifier).lower() == str(name_cell).lower():
            ws.delete_rows(row[0].row, 1)
            removed = True
            break

    wb.save(filepath)
    return removed
