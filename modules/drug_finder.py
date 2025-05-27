import customtkinter as ctk
from openpyxl import load_workbook
import os

class drugs(ctk.CTkToplevel):
    def __init__(self, master=None):
        super().__init__(master)
        self.title("Przeglądaj leki")
        self.geometry("800x700")
        self.configure(fg_color="#1f2937")
        self.focus()
        self.grab_set()

        # Nagłówek wyśrodkowany
        ctk.CTkLabel(
            self,
            text="Lista leków",
            font=("Arial", 28, "bold"),
            text_color="#329e76"
        ).pack(pady=(20, 10))

        # Scrollowalna ramka
        self.scroll_frame = ctk.CTkScrollableFrame(
            self,
            fg_color="#1f2937",
            width=750,
            height=550
        )
        self.scroll_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self.load_drugs_from_excel()

    def load_drugs_from_excel(self):
        filepath = os.path.join("database", "drugs.xlsx")
        workbook = load_workbook(filepath)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            id, nazwa, cena, stan_magazynowy, *_ = row

            card = ctk.CTkFrame(
                self.scroll_frame,
                fg_color="#2c3e50",
                height=120
            )
            card.pack(pady=10, fill="x", padx=10)

            ctk.CTkLabel(
                card,
                text=nazwa,
                font=("Arial", 18, "bold"),
                text_color="#329e76"
            ).pack(anchor="w", padx=10, pady=(5, 0))

            info_frame = ctk.CTkFrame(card, fg_color="transparent")
            info_frame.pack(fill="x", padx=10, pady=5)

            ctk.CTkLabel(
                info_frame,
                text=f"ID: {id}",
                font=("Arial", 14),
                text_color="#dce2e2"
            ).pack(side="left", padx=5)

            ctk.CTkLabel(
                info_frame,
                text=f"Cena: {cena} zł",
                font=("Arial", 14),
                text_color="#dce2e2"
            ).pack(side="left", padx=15)

            ctk.CTkLabel(
                info_frame,
                text=f"Stan: {int(stan_magazynowy)}",
                font=("Arial", 14),
                text_color="#dce2e2"
            ).pack(side="left", padx=5)